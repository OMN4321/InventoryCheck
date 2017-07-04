'''
Created on 1 Oct 2015

@author: Emanuele
'''

from openpyxl import load_workbook, Workbook
from openpyxl.utils.exceptions import InvalidFileException
from openpyxl.utils import get_column_letter
from openpyxl.styles import colors, PatternFill
from openpyxl.styles.fonts import Font
from openpyxl.styles.borders import Border, Side


def getData(filePath, fileType="storage"):
    '''
    open the file from filePath and
    return partNumberList, descriptionList, quantityList
    '''
    if fileType == "storage":
        sheet = 'STORAGE'
        pnColumn = 1
        descriptionColumn = 3
        quantityColumn = 4
    else:
        sheet = 'Product Service List'
        pnColumn = 2
        descriptionColumn = 4 
        quantityColumn = 7
    
    try:
        wb = load_workbook(filePath)
        ws = wb[sheet]
    except InvalidFileException:
        print("The file",filePath,"has a wrong format. Please save this file in order to have the xlsx format")
    except FileNotFoundError:
        print("The file",filePath,"is not found")
    except KeyError:
        print("Make sure the name of the sheets are: STORAGE and Worksheet")
    except:
        print("For some reason the file",filePath,"cannot be opened")
        print("Make sure the name of the files are: Storage.xlsx and ProductServiceList.xls")
        raise
    
    partNumberList = []
    descriptionList = []
    quantityList = []

    i = 2
    while(i in range(2,ws.max_row +1)):
        pn = str(ws.cell(row = i, column = pnColumn).value)
        description = str(ws.cell(row = i, column = descriptionColumn).value)
        try:
            qty = int(ws.cell(row = i, column = quantityColumn).value)
        except (TypeError,ValueError):
            qty = ws.cell(row = i, column = quantityColumn).value
            if qty is None:
                qty = 0
                
             
        if pn != '' and pn != 'product/service' and qty != '':    
            partNumberList.append(pn.lower())
            descriptionList.append(description)
            quantityList.append(qty)
            i=i+1
        else:
            i=i+1
            
    return (partNumberList, descriptionList, quantityList)

def createFile(filePath):
    '''
    create an excel file in the filePath
    '''
    
    wb = Workbook()
    ws = wb.create_sheet('Report',0)

    header = ['P/N', 'Description', 'Storage-Qty', 'Quickbook-Qty', 'Real-Qty']
    j = 1
    for i in header:
        ws.cell(row = 1, column = j, value = i).font = Font(color = colors.BLACK, sz=13, bold = True)       
        j=j+1
    
    wb.save(filePath)
    
    return wb

    
def writeInto(filePath,partNum,description,storageQty,quickbookQty):
    '''
    write all the data (partNum, Description, storageQty, QuickbookQty) 
    into an excel file in filePath
    '''
    wb = load_workbook(filePath)
    ws=wb['Report']
    i = 0
    j = 0
    while(i in range(0,len(partNum))):
        j=i+2 
        ws.cell(row = j, column = 1, value = partNum[i]) 
        ws.cell(row = j, column = 2, value = description[i])
        ws.cell(row = j, column = 3, value = storageQty[i])
        ws.cell(row = j, column = 4, value = quickbookQty[i])
        i=i+1
        
        
    wb.save(filePath)

def columnAdjustment(filePath):
    '''
    adjust the width of the column in base on the biggest string
    '''
    missmatch = False
    wb = load_workbook(filePath)
    ws = wb['Report']
    j=1
    i=1
    for j in range(1,6):
        for i in range(1,ws.max_row+1):
            cell = ws.cell(row = i, column = j)
            cell.alignment = cell.alignment.copy(horizontal='center', vertical='center')
            border = Border(bottom = Side(style='thin', color="FF000000"), left = Side(style='thin', color="FF000000"))
            cell.border = border
            if ws.cell(row = i, column = 3).value != ws.cell(row = i, column = 4).value and i != 1:
                missmatch = True
                cell.fill = PatternFill(start_color='caccce', end_color='caccce', fill_type='solid')
    width = 0
    widthNew = 0
    j=1
    i=1
    for j in range(1,3):
        for i in range(1,ws.max_row+1):
            widthNew=len(str(ws.cell(row = i, column = j).value))
            if widthNew > width:
                width = widthNew
                
        ws.column_dimensions[get_column_letter(j)].width = width+2
    ws.column_dimensions[get_column_letter(3)].width = len(str(ws.cell(row = 1, column = 3).value))+5
    ws.column_dimensions[get_column_letter(4)].width = len(str(ws.cell(row = 1, column = 4).value))+5
    
    
    wb.save(filePath)   
    return missmatch

